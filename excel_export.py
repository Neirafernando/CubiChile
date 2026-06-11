from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from database import get_project_by_id, get_presupuesto_by_project, get_budget_total


def format_money(value):
    return float(value or 0)


def unidad_visible(unidad):
    unidades = {
        "m2": "m²",
        "m3": "m³",
        "kg": "kg",
        "ml": "ml",
        "un": "un"
    }
    return unidades.get(unidad, unidad or "")


def export_presupuesto_excel(project_id, filepath):
    project = get_project_by_id(project_id)
    items = get_presupuesto_by_project(project_id)
    total_general = get_budget_total(project_id)

    project_id, name, client, location, description, created_at = project

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    # Colores
    dark = "0F172A"
    blue = "2563EB"
    light = "F8FAFC"
    gray = "E2E8F0"
    border_color = "CBD5E1"

    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"] = "CubiChile - Presupuesto de Obra"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Datos proyecto
    ws["A3"] = "Proyecto"
    ws["B3"] = name

    ws["A4"] = "Cliente"
    ws["B4"] = client or "Sin cliente"

    ws["A5"] = "Ubicación"
    ws["B5"] = location or "Sin ubicación"

    ws["A6"] = "Fecha creación"
    ws["B6"] = created_at

    for cell in ["A3", "A4", "A5", "A6"]:
        ws[cell].font = Font(bold=True, color="FFFFFF")
        ws[cell].fill = PatternFill("solid", fgColor=blue)
        ws[cell].border = border

    for cell in ["B3", "B4", "B5", "B6"]:
        ws[cell].fill = PatternFill("solid", fgColor=light)
        ws[cell].border = border

    # Encabezados tabla
    headers = [
        "Ítem",
        "Partida",
        "Cantidad",
        "Unidad",
        "Precio unitario",
        "Total",
        "Fecha"
    ]

    start_row = 8

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    row = start_row + 1

    for index, item in enumerate(items, 1):
        cub_id, tipo, volumen, unidad, precio_unitario, total, item_created_at = item

        ws.cell(row=row, column=1).value = index
        ws.cell(row=row, column=2).value = tipo
        ws.cell(row=row, column=3).value = float(volumen or 0)
        ws.cell(row=row, column=4).value = unidad_visible(unidad or "m3")
        ws.cell(row=row, column=5).value = format_money(precio_unitario)
        ws.cell(row=row, column=6).value = format_money(total)
        ws.cell(row=row, column=7).value = item_created_at

        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(vertical="center")

        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=3).number_format = '0.00'

        row += 1

    # Total
    total_row = row + 1

    ws.cell(row=total_row, column=5).value = "TOTAL PRESUPUESTO"
    ws.cell(row=total_row, column=5).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=5).fill = PatternFill("solid", fgColor=blue)
    ws.cell(row=total_row, column=5).border = border

    ws.cell(row=total_row, column=6).value = format_money(total_general)
    ws.cell(row=total_row, column=6).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=6).fill = PatternFill("solid", fgColor=blue)
    ws.cell(row=total_row, column=6).number_format = '$#,##0'
    ws.cell(row=total_row, column=6).border = border

    # Anchos
    widths = {
        "A": 8,
        "B": 34,
        "C": 14,
        "D": 12,
        "E": 18,
        "F": 18,
        "G": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Alturas
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[start_row].height = 22

    # Congelar encabezado
    ws.freeze_panes = "A9"

    wb.save(filepath)
